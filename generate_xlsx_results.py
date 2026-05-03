"""
Generate XLSX evaluation report from experiment results.

Sheet 1 ("Detailed"): Three tables (Completion %, Avg Time, Avg Steps)
  - Rows: Model x Modality
  - Columns: 7 process adaptations (base experiment) + 2 exception columns + Avg
  - Grouped: Operational | Tactical | Exceptions

Sheet 2 ("Summary"): One table
  - Rows: Model x Modality
  - Columns: Overall Completion %, Time avg +/- std, Steps avg +/- std
"""

import json
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "experiment_results"
SESSIONS_DIR = BASE_DIR / "sessions"
OUTPUT_DIR = BASE_DIR / "results"

# --- Column definitions ---

# Operational adaptations (base experiment type)
OPERATIONAL = [
    ("base_rule", "Base Rule"),
    ("0_values", "Area 0"),
    ("500_values", "Area 500"),
    ("900_values", "Area 900"),
    ("city_values", "Excl. Cities"),
]

# Tactical adaptations (base experiment type)
TACTICAL = [
    ("extension_estimates", "Mand. Reading"),
    ("extension_mail", "Dir. Mail"),
]

# Exception experiment types (aggregated across all adaptations)
EXCEPTIONS = [
    ("exception_handling", "Typing Err."),
    ("exception_handling_db_error", "DB Err."),
]

ALL_BASE_ADAPTATIONS = [key for key, _ in OPERATIONAL + TACTICAL]

# --- Method config ---

METHOD_CONFIG = {
    "add": {
        "display": "Modular Agents Add Rule",
        "json_files": [
            "frame_agent_process_addition.json",
            "frame_agent_process_execution.json",
            "process_agent.json",
        ],
    },
    "generate_bpmn": {
        "display": "Modular Agents BPMN Rule",
        "json_files": [
            "frame_agent_process_addition.json",
            "frame_agent_process_execution.json",
            "process_agent.json",
        ],
    },
    "classic": {
        "display": "Classic Agent Prompt",
        "json_files": [
            "classic_agent.json",
        ],
    },
}

MODEL_DISPLAY = {
    "gpt-5.4": "GPT-5.4",
    "gpt-5.1": "GPT-5.1",
    "qwen3.5_35b": "Qwen3.5:35B",
}

# Display order for models in tables and summary output.
MODEL_ORDER = ["gpt-5.4", "gpt-5.1", "qwen3.5_35b"]

# Reasoning effort values, in display order. None represents "no reasoning component"
# (Ollama models, or legacy CSVs/sessions written before reasoning_effort was tracked).
REASONING_VALUES = ["none", "minimal", "low", "medium", "high"]

# Methods ordered longest-first so name parsing finds the right anchor
# (`generate_bpmn` must be tested before `add` to avoid a false-positive match).
METHODS_ORDERED = ["generate_bpmn", "classic", "add"]

CSV_PREFIX = "process_adaptation_results_"


@dataclass(frozen=True)
class Variant:
    """One (model, reasoning_effort) combination — the unit of comparison in result tables."""
    model: str
    reasoning: Optional[str]  # None means no reasoning component (e.g. Ollama, legacy)

    @property
    def key(self) -> str:
        return self.model if self.reasoning is None else f"{self.model}_reasoning-{self.reasoning}"

    @property
    def display(self) -> str:
        base = MODEL_DISPLAY.get(self.model, self.model)
        if self.reasoning is None:
            return base
        return f"{base} (reasoning={self.reasoning})"

    @property
    def model_display(self) -> str:
        return MODEL_DISPLAY.get(self.model, self.model)

    @property
    def reasoning_display(self) -> str:
        return self.reasoning if self.reasoning is not None else "—"

    @property
    def session_dir(self) -> Path:
        if self.reasoning is None:
            return SESSIONS_DIR / self.model
        return SESSIONS_DIR / self.model / f"reasoning-{self.reasoning}"

    @property
    def csv_prefix(self) -> str:
        if self.reasoning is None:
            return f"{CSV_PREFIX}{self.model}_"
        return f"{CSV_PREFIX}{self.model}_reasoning-{self.reasoning}_"


# --- Data loading helpers ---


def _parse_csv_stem(stem: str) -> Optional[tuple[str, Optional[str], str, str]]:
    """Parse a result-CSV stem into (model, reasoning, method, experiment_type).

    Accepted patterns:
      process_adaptation_results_<model>_<method>[<exception_suffix>]
      process_adaptation_results_<model>_reasoning-<value>_<method>[<exception_suffix>]
    Returns None if the stem does not match either pattern.
    """
    if not stem.startswith(CSV_PREFIX):
        return None
    rest = stem[len(CSV_PREFIX):]

    method = None
    method_idx = -1
    for m in METHODS_ORDERED:
        marker = f"_{m}"
        idx = rest.find(marker)
        if idx > 0 and (idx + len(marker) == len(rest) or rest[idx + len(marker)] == "_"):
            method = m
            method_idx = idx
            break
    if method is None:
        return None

    model_part = rest[:method_idx]
    suffix_part = rest[method_idx + 1 + len(method):]  # everything after `_<method>`

    reasoning: Optional[str] = None
    model = model_part
    for r in REASONING_VALUES:
        marker = f"_reasoning-{r}"
        if model_part.endswith(marker):
            reasoning = r
            model = model_part[: -len(marker)]
            break

    if suffix_part == "":
        experiment_type = "base"
    elif suffix_part == "_exception_handling":
        experiment_type = "exception_handling"
    elif suffix_part == "_exception_handling_db_error":
        experiment_type = "exception_handling_db_error"
    else:
        experiment_type = "base"

    return model, reasoning, method, experiment_type


def discover_variants() -> list[Variant]:
    """Find all (model, reasoning) variants represented by result CSVs.

    Variants are ordered by MODEL_ORDER, then by REASONING_VALUES (None first).
    Unknown models are appended alphabetically.
    """
    found: set[tuple[str, Optional[str]]] = set()
    for f in RESULTS_DIR.iterdir():
        if not (f.is_file() and f.suffix == ".csv"):
            continue
        parsed = _parse_csv_stem(f.stem)
        if parsed is None:
            continue
        model, reasoning, _method, _exp = parsed
        found.add((model, reasoning))

    def sort_key(v: tuple[str, Optional[str]]) -> tuple:
        model, reasoning = v
        if model in MODEL_ORDER:
            mi = MODEL_ORDER.index(model)
        else:
            mi = len(MODEL_ORDER)
        if reasoning is None:
            ri = -1
        elif reasoning in REASONING_VALUES:
            ri = REASONING_VALUES.index(reasoning)
        else:
            ri = len(REASONING_VALUES)
        return (mi, ri, model, reasoning or "")

    return [Variant(model=m, reasoning=r) for m, r in sorted(found, key=sort_key)]


def load_result_csvs(variant: Variant) -> pd.DataFrame:
    """Load and concatenate all result CSVs that belong to this variant."""
    frames = []
    for f in sorted(RESULTS_DIR.iterdir()):
        if not (f.is_file() and f.suffix == ".csv"):
            continue
        parsed = _parse_csv_stem(f.stem)
        if parsed is None:
            continue
        model, reasoning, _method, experiment_type = parsed
        if model != variant.model or reasoning != variant.reasoning:
            continue
        df = pd.read_csv(f)
        df["experiment_type"] = experiment_type
        frames.append(df)

    if not frames:
        raise FileNotFoundError(f"No result CSVs found for variant '{variant.key}'")
    return pd.concat(frames, ignore_index=True)


def read_timings(session_path: Path) -> dict:
    timings_file = session_path / "timings.csv"
    result = {}
    if timings_file.exists():
        with open(timings_file) as f:
            for row in csv.DictReader(f):
                result[row["agent"]] = float(row["time"])
    return result


def get_total_time(timings: dict, method: str) -> float | None:
    if method in ("add", "generate_bpmn"):
        f, p = timings.get("frame"), timings.get("process")
        return (f + p) if f is not None and p is not None else None
    return timings.get("classic")


def count_ai_messages(session_path: Path, method: str) -> int | None:
    total = 0
    for jf in METHOD_CONFIG[method]["json_files"]:
        jpath = session_path / jf
        if not jpath.exists():
            return None
        try:
            with open(jpath) as f:
                data = json.load(f)
            msgs = data.get("messages", data) if isinstance(data, dict) else data
            total += sum(
                1 for m in msgs
                if isinstance(m, dict) and m.get("id", [None])[-1] == "AIMessage"
            )
        except (json.JSONDecodeError, KeyError):
            return None
    return total


def enrich_with_session_data(df: pd.DataFrame, variant: Variant) -> pd.DataFrame:
    """Add total_time_s and ai_steps columns from session files."""
    times, ai_steps = [], []
    base_dir = variant.session_dir
    for _, row in df.iterrows():
        sp = base_dir / row["session_id"]
        method = row["rule_adaptation_method"]
        timings = read_timings(sp)
        times.append(get_total_time(timings, method))
        ai_steps.append(count_ai_messages(sp, method))
    df = df.copy()
    df["total_time_s"] = times
    df["ai_steps"] = ai_steps
    return df


# --- Report building ---


def _completion_pct(subset: pd.DataFrame) -> float | None:
    if len(subset) == 0:
        return None
    return round(subset["all_correct"].sum() / len(subset) * 100, 1)


def _mean_or_none(series: pd.Series) -> float | None:
    valid = series.dropna()
    return round(valid.mean(), 1) if len(valid) > 0 else None


def _std_or_none(series: pd.Series) -> float | None:
    valid = series.dropna()
    return round(valid.std(), 1) if len(valid) > 1 else None


def _fmt_mean_std(mean_val, std_val) -> str | None:
    if mean_val is None:
        return None
    if std_val is None or np.isnan(std_val):
        return f"{mean_val}"
    return f"{mean_val} \u00b1 {std_val}"


def build_detailed_table(df: pd.DataFrame, variant: Variant, metric: str) -> list[dict]:
    """Build rows for one metric table on the Detailed sheet.
    metric: 'completion', 'time', or 'steps'
    """
    rows = []

    for method, cfg in METHOD_CONFIG.items():
        method_df = df[df["rule_adaptation_method"] == method]
        if method_df.empty:
            continue

        row = {
            "Model": variant.model_display,
            "Reasoning": variant.reasoning_display,
            "Modality": cfg["display"],
        }

        # Operational + Tactical columns: base experiment type, per adaptation
        values_for_avg = []
        for adapt_key, adapt_label in OPERATIONAL + TACTICAL:
            subset = method_df[
                (method_df["process_adaptation"] == adapt_key)
                & (method_df["experiment_type"] == "base")
            ]
            if metric == "completion":
                val = _completion_pct(subset)
            elif metric == "time":
                val = _mean_or_none(subset["total_time_s"])
            else:  # steps
                val = _mean_or_none(subset["ai_steps"])
            row[adapt_label] = val
            if val is not None:
                values_for_avg.append(val)

        # Exception columns: aggregated across all adaptations for that experiment type
        for exp_key, exp_label in EXCEPTIONS:
            subset = method_df[method_df["experiment_type"] == exp_key]
            if metric == "completion":
                val = _completion_pct(subset)
            elif metric == "time":
                val = _mean_or_none(subset["total_time_s"])
            else:
                val = _mean_or_none(subset["ai_steps"])
            row[exp_label] = val
            if val is not None:
                values_for_avg.append(val)

        # Avg column
        row["Avg."] = round(np.mean(values_for_avg), 1) if values_for_avg else None
        rows.append(row)

    return rows


def build_summary_table(df: pd.DataFrame, variant: Variant) -> list[dict]:
    """Build rows for the Summary sheet."""
    rows = []

    for method, cfg in METHOD_CONFIG.items():
        method_df = df[df["rule_adaptation_method"] == method]
        if method_df.empty:
            continue

        row = {
            "Model": variant.model_display,
            "Reasoning": variant.reasoning_display,
            "Modality": cfg["display"],
        }

        # Overall completion across ALL experiment types and adaptations
        total = len(method_df)
        correct = method_df["all_correct"].sum()
        row["Completion (%)"] = round(correct / total * 100, 1) if total > 0 else None

        # Time
        t_mean = _mean_or_none(method_df["total_time_s"])
        t_std = _std_or_none(method_df["total_time_s"])
        row["Avg Time (s)"] = _fmt_mean_std(t_mean, t_std)

        # Steps
        s_mean = _mean_or_none(method_df["ai_steps"])
        s_std = _std_or_none(method_df["ai_steps"])
        row["Avg Steps"] = _fmt_mean_std(s_mean, s_std)

        rows.append(row)

    return rows


# --- XLSX formatting ---

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
GROUP_FONT = Font(bold=True, size=10)
TABLE_TITLE_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GROUP_RIGHT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="medium"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def auto_width(ws, df: pd.DataFrame, col_offset: int = 0):
    """Set column widths based on content."""
    for col_idx, col in enumerate(df.columns, 1):
        values = [len(str(v)) for v in df.iloc[:, col_idx - 1] if pd.notna(v)]
        max_len = max(len(str(col)), *values) if values else len(str(col))
        letter = get_column_letter(col_idx + col_offset)
        ws.column_dimensions[letter].width = max_len + 3


def _apply_bold_underline(ws, data_rows: list[dict], columns: list[str],
                          start_row: int, col_offset: int):
    """Bold the best value and underline second best per column (for completion: higher is better)."""
    for col_name in columns:
        col_idx = list(data_rows[0].keys()).index(col_name) + 1 + col_offset
        vals = []
        for i, row in enumerate(data_rows):
            v = row.get(col_name)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                vals.append((v, i))

        if len(vals) < 2:
            continue

        vals.sort(key=lambda x: x[0], reverse=True)
        best_val, best_idx = vals[0]
        second_val, second_idx = vals[1]

        # Bold best
        cell = ws.cell(row=start_row + best_idx, column=col_idx)
        cell.font = Font(bold=True, size=10)

        # Underline second best
        cell = ws.cell(row=start_row + second_idx, column=col_idx)
        cell.font = Font(underline="single", size=10)


def write_detailed_sheet(ws, all_variants_data: dict):
    """Write three tables (Completion, Time, Steps) vertically on the Detailed sheet."""

    metric_configs = [
        ("completion", "Table 3: Completion Rate (%)"),
        ("time", "Table A1: Average Time (s)"),
        ("steps", "Table A2: Average Steps (AI Messages)"),
    ]

    # Column structure for group headers
    # Cols A/B/C = Model / Reasoning / Modality; then Operational (5), Tactical (2), Exceptions (2), Avg (1)
    group_spans = [
        ("Operational", 4, 8),      # columns D-H (5 cols)
        ("Tactical", 9, 10),        # columns I-J (2 cols)
        ("Exceptions", 11, 12),     # columns K-L (2 cols)
    ]

    current_row = 1

    for metric, title in metric_configs:
        # Table title
        ws.cell(row=current_row, column=1, value=title).font = TABLE_TITLE_FONT
        current_row += 1

        # Group header row
        ws.cell(row=current_row, column=1, value="")
        for group_name, start_col, end_col in group_spans:
            ws.merge_cells(
                start_row=current_row, start_column=start_col,
                end_row=current_row, end_column=end_col
            )
            cell = ws.cell(row=current_row, column=start_col, value=group_name)
            cell.font = GROUP_FONT
            cell.fill = GROUP_FILL
            cell.alignment = Alignment(horizontal="center")
            # Fill all merged cells
            for c in range(start_col, end_col + 1):
                ws.cell(row=current_row, column=c).fill = GROUP_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Collect rows from all variants
        all_rows = []
        for variant, df in all_variants_data.items():
            all_rows.extend(build_detailed_table(df, variant, metric))

        if not all_rows:
            current_row += 1
            continue

        # Column headers
        columns = list(all_rows[0].keys())
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        current_row += 1

        # Data rows
        data_start_row = current_row
        for row_data in all_rows:
            for col_idx, col_name in enumerate(columns, 1):
                val = row_data.get(col_name)
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                # Left-align the 3 leading text columns (Model / Reasoning / Modality);
                # center-align all numeric/data columns.
                cell.alignment = Alignment(horizontal="center") if col_idx > 3 else Alignment()
                # Medium border at group boundaries (right edge of Operational/Tactical/Exceptions)
                if col_idx in (8, 10, 12):
                    cell.border = GROUP_RIGHT_BORDER
            current_row += 1

        # Bold best / underline second best — skip the 3 leading header columns
        value_columns = columns[3:]
        _apply_bold_underline(ws, all_rows, value_columns, data_start_row, col_offset=0)

        # Blank row between tables
        current_row += 2

    # Column widths
    col_widths = {"A": 14, "B": 12, "C": 28}
    for i in range(4, 14):
        col_widths[get_column_letter(i)] = 14
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "D4"


def write_summary_sheet(ws, all_variants_data: dict):
    """Write the summary table on the Summary sheet."""
    ws.cell(row=1, column=1, value="Table 4: Summary across Models and Modalities").font = TABLE_TITLE_FONT

    all_rows = []
    for variant, df in all_variants_data.items():
        all_rows.extend(build_summary_table(df, variant))

    if not all_rows:
        return

    columns = list(all_rows[0].keys())

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data
    data_start_row = 3
    for row_idx, row_data in enumerate(all_rows):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=data_start_row + row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            # Left-align Model / Reasoning / Modality; center-align metric columns
            cell.alignment = Alignment(horizontal="center") if col_idx > 3 else Alignment()

    # Bold best / underline second best for Completion column
    _apply_bold_underline(ws, all_rows, ["Completion (%)"], data_start_row, col_offset=0)

    # Column widths: Model / Reasoning / Modality, then metric columns
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    for i in range(4, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "D3"


# --- Main ---


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    variants = discover_variants()

    if not variants:
        print("No result CSVs found in experiment_results/")
        return

    # Load and enrich data for all variants
    all_variants_data: dict[Variant, pd.DataFrame] = {}
    for variant in variants:
        print(f"Loading {variant.display} ({variant.key})...")
        try:
            df = load_result_csvs(variant)
            df = enrich_with_session_data(df, variant)
            all_variants_data[variant] = df
        except FileNotFoundError as e:
            print(f"  Skipping: {e}")

    # Write one XLSX per variant
    for variant, df in all_variants_data.items():
        single = {variant: df}
        output_path = OUTPUT_DIR / f"{variant.key}_results.xlsx"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Create placeholder sheets so we can access the workbook
            pd.DataFrame().to_excel(writer, sheet_name="Detailed", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
            wb = writer.book

            ws_detail = wb["Detailed"]
            write_detailed_sheet(ws_detail, single)

            ws_summary = wb["Summary"]
            write_summary_sheet(ws_summary, single)

        print(f"  Written: {output_path}")

    # Combined XLSX with all variants
    if len(all_variants_data) > 1:
        output_path = OUTPUT_DIR / "all_models_results.xlsx"
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Detailed", index=False)
            pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
            wb = writer.book

            ws_detail = wb["Detailed"]
            write_detailed_sheet(ws_detail, all_variants_data)

            ws_summary = wb["Summary"]
            write_summary_sheet(ws_summary, all_variants_data)

        print(f"  Written: {output_path}")

    # Print summary to console
    print("\n--- Summary ---")
    for variant, df in all_variants_data.items():
        for method, cfg in METHOD_CONFIG.items():
            mdf = df[df["rule_adaptation_method"] == method]
            if mdf.empty:
                continue
            total = len(mdf)
            correct = mdf["all_correct"].sum()
            t = mdf["total_time_s"].dropna()
            s = mdf["ai_steps"].dropna()
            print(
                f"{variant.display} {cfg['display']:30s}  "
                f"Completion: {correct/total*100:5.1f}%  "
                f"Time: {t.mean():6.1f} \u00b1 {t.std():5.1f}s  "
                f"Steps: {s.mean():5.1f} \u00b1 {s.std():4.1f}"
            )


if __name__ == "__main__":
    main()
